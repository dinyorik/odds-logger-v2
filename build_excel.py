"""
Создаёт odds_dashboard.xlsx с пятью листами на основе odds_log.csv и matches.csv.

Запуск:
    python build_excel.py

Excel перед запуском надо закрыть (иначе ошибка "файл занят").
"""

import csv
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series

from paths import LOG_FILE
from paths import MATCHES_FILE
from paths import HISTORICAL_FILE
from paths import DASHBOARD_FILE as OUT_FILE

WINDOW_THRESHOLD = 1.30
COMEBACK_MAX_THRESHOLD = 5.0  # лист Comeback hunt: только матчи где max_cf проигравшего >= 5


# ---------------- helpers ----------------

def read_csv(path):
    if not os.path.exists(path):
        return [], []
    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def autosize(ws, ncols, min_w=10, max_w=22):
    for c in range(1, ncols + 1):
        col_letter = get_column_letter(c)
        max_len = min_w
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=True):
            v = row[0]
            if v is not None:
                max_len = max(max_len, min(len(str(v)) + 2, max_w))
        ws.column_dimensions[col_letter].width = max_len


def to_float(s):
    try:
        return float(s) if s not in ("", None) else None
    except (ValueError, TypeError):
        return None


# ---------------- sheet 1: tick log ----------------

def build_log_sheet(ws, header, rows):
    if not header:
        ws["A1"] = "odds_log.csv не найден или пустой — запусти сначала odds_logger.py"
        return

    ws.append(header)
    style_header(ws, len(header))

    for r in rows:
        out = list(r)
        for col_name in ("cf_team_a", "cf_team_b", "min_cf_so_far", "map", "sec_since_start"):
            if col_name in header:
                idx = header.index(col_name)
                if idx < len(out):
                    f = to_float(out[idx])
                    if f is not None:
                        out[idx] = f
        ws.append(out)

    n = len(rows) + 1

    if "min_cf_so_far" in header and "blocked" in header:
        col_min = get_column_letter(header.index("min_cf_so_far") + 1)
        col_blocked = get_column_letter(header.index("blocked") + 1)
        rng = f"A2:{get_column_letter(len(header))}{n}"
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'AND(${col_min}2<={WINDOW_THRESHOLD},${col_blocked}2="False")'],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'${col_blocked}2="True"'],
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )

    autosize(ws, len(header))


# ---------------- sheet 2: matches summary ----------------

def build_matches_sheet(ws, header, rows):
    if not header:
        ws["A1"] = "matches.csv пока пустой — нужно дождаться окончания хотя бы одного матча"
        return

    extra = [
        "duration_sec",
        "window_duration_sec",
        "min_cf_winner",
        "max_cf_loser",
        "would_bet_win",
    ]
    full_header = header + extra
    ws.append(full_header)
    style_header(ws, len(full_header))

    h = {name: i for i, name in enumerate(header)}

    def get(row, col):
        if col in h and h[col] < len(row):
            return row[h[col]]
        return ""

    for r in rows:
        out = list(r)

        # numeric conversion
        for col in ("min_cf_a", "min_cf_b", "max_cf_a", "max_cf_b",
                    "last_open_cf_a", "last_open_cf_b",
                    "first_cf_le_threshold_a", "first_cf_le_threshold_b",
                    "ticks_logged", "map"):
            if col in h and h[col] < len(out):
                f = to_float(out[h[col]])
                if f is not None:
                    out[h[col]] = f

        # duration
        try:
            t1 = datetime.strptime(get(r, "started_at"), "%Y-%m-%d %H:%M:%S")
            t2 = datetime.strptime(get(r, "ended_at"), "%Y-%m-%d %H:%M:%S")
            duration = int((t2 - t1).total_seconds())
        except Exception:
            duration = ""

        # window duration for the team that won
        winner = get(r, "winner")
        win_dur = ""
        try:
            if winner == "A" and get(r, "first_cf_le_threshold_a_at"):
                ta = datetime.strptime(get(r, "first_cf_le_threshold_a_at"), "%Y-%m-%d %H:%M:%S")
                tend = datetime.strptime(get(r, "ended_at"), "%Y-%m-%d %H:%M:%S")
                win_dur = int((tend - ta).total_seconds())
            elif winner == "B" and get(r, "first_cf_le_threshold_b_at"):
                tb = datetime.strptime(get(r, "first_cf_le_threshold_b_at"), "%Y-%m-%d %H:%M:%S")
                tend = datetime.strptime(get(r, "ended_at"), "%Y-%m-%d %H:%M:%S")
                win_dur = int((tend - tb).total_seconds())
        except Exception:
            pass

        min_cf_winner = ""
        max_cf_loser = ""
        if winner == "A":
            min_cf_winner = to_float(get(r, "min_cf_a")) or ""
            max_cf_loser = to_float(get(r, "max_cf_b")) or ""
        elif winner == "B":
            min_cf_winner = to_float(get(r, "min_cf_b")) or ""
            max_cf_loser = to_float(get(r, "max_cf_a")) or ""

        would_win = ""
        if winner == "A":
            would_win = "YES" if get(r, "first_cf_le_threshold_a") not in ("", None) else "NO"
        elif winner == "B":
            would_win = "YES" if get(r, "first_cf_le_threshold_b") not in ("", None) else "NO"

        ws.append(out + [duration, win_dur, min_cf_winner, max_cf_loser, would_win])

    n = len(rows) + 1

    if len(rows) > 0:
        win_col = get_column_letter(full_header.index("would_bet_win") + 1)
        ws.conditional_formatting.add(
            f"{win_col}2:{win_col}{n}",
            FormulaRule(formula=[f'{win_col}2="YES"'],
                        fill=PatternFill("solid", fgColor="C6EFCE")),
        )
        ws.conditional_formatting.add(
            f"{win_col}2:{win_col}{n}",
            FormulaRule(formula=[f'{win_col}2="NO"'],
                        fill=PatternFill("solid", fgColor="FFC7CE")),
        )

    autosize(ws, len(full_header))


# ---------------- sheet 3: stats by bin ----------------

def build_stats_sheet(ws, matches_header, matches_rows):
    if not matches_header:
        ws["A1"] = "Нет данных в matches.csv"
        return

    h = {name: i for i, name in enumerate(matches_header)}

    bets = []
    for r in matches_rows:
        winner = r[h["winner"]] if h["winner"] < len(r) else ""
        if winner not in ("A", "B"):
            continue

        a_first = to_float(r[h["first_cf_le_threshold_a"]]) if h["first_cf_le_threshold_a"] < len(r) else None
        b_first = to_float(r[h["first_cf_le_threshold_b"]]) if h["first_cf_le_threshold_b"] < len(r) else None
        a_first_at = r[h["first_cf_le_threshold_a_at"]] if h["first_cf_le_threshold_a_at"] < len(r) else ""
        b_first_at = r[h["first_cf_le_threshold_b_at"]] if h["first_cf_le_threshold_b_at"] < len(r) else ""

        bet_on = None
        bet_cf = None
        if a_first is not None and b_first is not None:
            bet_on = "A" if a_first_at <= b_first_at else "B"
            bet_cf = a_first if bet_on == "A" else b_first
        elif a_first is not None:
            bet_on, bet_cf = "A", a_first
        elif b_first is not None:
            bet_on, bet_cf = "B", b_first

        if bet_on is None:
            continue

        won = (bet_on == winner)
        bets.append((bet_cf, won))

    bins = [
        ("<1.05", 0.0, 1.05),
        ("1.05-1.10", 1.05, 1.10),
        ("1.10-1.15", 1.10, 1.15),
        ("1.15-1.20", 1.15, 1.20),
        ("1.20-1.25", 1.20, 1.25),
        ("1.25-1.30", 1.25, 1.30),
        (">=1.30", 1.30, 99.0),
    ]

    ws.append(["Бин по first_cf (кэф входа)", "Ставок", "Побед", "Поражений",
               "Винрейт %", "Средний кэф входа", "EV %",
               "Kelly fraction %", "Комментарий"])
    style_header(ws, 9)

    total_w = total_l = 0
    for label, lo, hi in bins:
        in_bin = [b for b in bets if lo <= b[0] < hi]
        n = len(in_bin)
        wins = sum(1 for b in in_bin if b[1])
        losses = n - wins
        wr = wins / n if n else 0
        avg_cf = sum(b[0] for b in in_bin) / n if n else 0
        ev = (wr * (avg_cf - 1) - (1 - wr)) * 100 if n else 0
        if n and avg_cf > 1:
            b_val = avg_cf - 1
            kelly = (wr * b_val - (1 - wr)) / b_val * 100
        else:
            kelly = 0
        comment = ""
        if n < 10:
            comment = "мало данных"
        elif ev < 0:
            comment = "EV отрицательный — не ставить"
        elif kelly > 0:
            comment = f"ставить ≈{kelly:.1f}% банка по Kelly"

        ws.append([
            label, n, wins, losses,
            round(wr * 100, 1) if n else "",
            round(avg_cf, 4) if n else "",
            round(ev, 2) if n else "",
            round(kelly, 2) if n else "",
            comment,
        ])
        total_w += wins
        total_l += losses

    total_n = total_w + total_l
    total_wr = total_w / total_n if total_n else 0
    ws.append([])
    ws.append(["ИТОГО", total_n, total_w, total_l,
               round(total_wr * 100, 1) if total_n else "", "", "", "", ""])
    last = ws.max_row
    for c in range(1, 10):
        ws.cell(row=last, column=c).font = Font(bold=True)

    autosize(ws, 9, min_w=14)


# ---------------- sheet 4: comeback hunt ----------------

def build_comeback_sheet(ws, matches_header, matches_rows):
    """
    Гипотеза: команда у которой кэф улетал в пике на 5+ — иногда камбекает.
    Если это случается чаще чем подразумевается кэфом (т.е. >1/cf_max),
    есть положительный EV ставить на underdog когда его кэф пробивает потолок.

    Лист содержит только матчи где max_cf >= COMEBACK_MAX_THRESHOLD у одной из сторон.
    """
    if not matches_header:
        ws["A1"] = "Нет данных"
        return

    h = {name: i for i, name in enumerate(matches_header)}

    headers = [
        "match_id", "map", "winner",
        "max_cf_a", "max_cf_a_at",
        "max_cf_b", "max_cf_b_at",
        "underdog",          # та сторона у кого max_cf был выше = underdog
        "underdog_max_cf",
        "underdog_won",       # "YES" если underdog в итоге выиграл
        "implied_prob_at_peak %",  # 1/max_cf — сколько бук давал underdog'у
        "would_have_paid_at_peak",  # сколько бы вернулось при ставке 1$ на underdog в пике
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    underdogs_total = 0
    underdogs_won = 0
    sum_implied_prob = 0.0

    for r in matches_rows:
        winner = r[h["winner"]] if h["winner"] < len(r) else ""
        max_a = to_float(r[h["max_cf_a"]]) if "max_cf_a" in h and h["max_cf_a"] < len(r) else None
        max_b = to_float(r[h["max_cf_b"]]) if "max_cf_b" in h and h["max_cf_b"] < len(r) else None

        if max_a is None and max_b is None:
            continue
        peak_max = max([x for x in (max_a, max_b) if x is not None])
        if peak_max < COMEBACK_MAX_THRESHOLD:
            continue

        # underdog = сторона с большим max_cf
        if max_a is not None and (max_b is None or max_a >= max_b):
            underdog = "A"
            udog_max = max_a
            udog_max_at = r[h["max_cf_a_at"]] if h["max_cf_a_at"] < len(r) else ""
        else:
            underdog = "B"
            udog_max = max_b
            udog_max_at = r[h["max_cf_b_at"]] if h["max_cf_b_at"] < len(r) else ""

        underdog_won = (winner == underdog)
        underdogs_total += 1
        if underdog_won:
            underdogs_won += 1

        implied = 1.0 / udog_max if udog_max else 0
        sum_implied_prob += implied

        ws.append([
            r[h["match_id"]] if h["match_id"] < len(r) else "",
            to_float(r[h["map"]]) if h["map"] < len(r) else "",
            winner,
            max_a if max_a is not None else "",
            r[h["max_cf_a_at"]] if h["max_cf_a_at"] < len(r) else "",
            max_b if max_b is not None else "",
            r[h["max_cf_b_at"]] if h["max_cf_b_at"] < len(r) else "",
            underdog,
            udog_max,
            "YES" if underdog_won else "NO",
            round(implied * 100, 2),
            round(udog_max, 2) if underdog_won else 0,
        ])

    n = ws.max_row

    # подсветка
    if n >= 2:
        wcol = get_column_letter(headers.index("underdog_won") + 1)
        ws.conditional_formatting.add(
            f"{wcol}2:{wcol}{n}",
            FormulaRule(formula=[f'{wcol}2="YES"'],
                        fill=PatternFill("solid", fgColor="C6EFCE")),
        )
        ws.conditional_formatting.add(
            f"{wcol}2:{wcol}{n}",
            FormulaRule(formula=[f'{wcol}2="NO"'],
                        fill=PatternFill("solid", fgColor="FFC7CE")),
        )

    # сводка снизу
    ws.append([])
    ws.append([f"Всего underdog'ов с пиком ≥ {COMEBACK_MAX_THRESHOLD}", underdogs_total])
    ws.append(["Из них выиграли (камбэк)", underdogs_won])
    actual_rate = underdogs_won / underdogs_total if underdogs_total else 0
    avg_implied = sum_implied_prob / underdogs_total if underdogs_total else 0
    ws.append(["Реальный винрейт underdog'а %", round(actual_rate * 100, 2)])
    ws.append(["Средняя подразумеваемая вероятность %", round(avg_implied * 100, 2)])

    edge_pct = (actual_rate - avg_implied) * 100
    ws.append(["Edge (реальная − подразумеваемая) %", round(edge_pct, 2)])
    if underdogs_total < 30:
        ws.append(["", "мало данных, нужен 30+ underdog'ов для надёжного вывода"])
    elif edge_pct > 0:
        ws.append(["", "ПОЛОЖИТЕЛЬНЫЙ ЭДЖ — стоит ставить на underdog'а в пике"])
    else:
        ws.append(["", "эджа нет"])

    for row in ws.iter_rows(min_row=n + 1):
        for cell in row:
            cell.font = Font(bold=True)

    autosize(ws, len(headers), min_w=14)


# ---------------- sheet 5: score vs odds scatter ----------------

def build_scatter_sheet(ws, log_header, log_rows):
    if not log_header or "score" not in log_header or "cf_team_a" not in log_header:
        ws["A1"] = "Нет данных для скаттера"
        return

    h = {name: i for i, name in enumerate(log_header)}
    ws.append(["match_id", "map", "kill_diff (A-B)", "cf_team_a", "cf_team_b", "min_cf"])
    style_header(ws, 6)

    for r in log_rows:
        try:
            score = r[h["score"]]
            if not score or score == "?-?" or "-" not in score:
                continue
            a, b = score.split("-")
            kd = int(a) - int(b)
            cfa = to_float(r[h["cf_team_a"]])
            cfb = to_float(r[h["cf_team_b"]])
            if cfa is None or cfb is None:
                continue
            min_cf = min(cfa, cfb)
            ws.append([r[h["match_id"]], to_float(r[h["map"]]), kd, cfa, cfb, min_cf])
        except (ValueError, IndexError):
            continue

    n = ws.max_row
    if n < 3:
        return

    chart = ScatterChart()
    chart.title = "Дельта килов vs min(cf_A, cf_B)"
    chart.x_axis.title = "Kill diff (A − B)"
    chart.y_axis.title = "min cf"
    chart.height = 12
    chart.width = 20

    x = Reference(ws, min_col=3, min_row=2, max_row=n)
    y = Reference(ws, min_col=6, min_row=2, max_row=n)
    chart.series.append(Series(y, x, title_from_data=False))
    chart.legend = None

    ws.add_chart(chart, "H2")
    autosize(ws, 6)


# ---------------- sheet 6: historical bets ----------------

def build_historical_sheet(ws, header, rows):
    if not header:
        ws["A1"] = "historical_bets.csv не найден — запусти parse_history.py"
        return
    ws.append(header)
    style_header(ws, len(header))
    for r in rows:
        out = list(r)
        for col_name in ("map", "odds", "stake_kgs", "payout_kgs", "profit_kgs"):
            if col_name in header:
                idx = header.index(col_name)
                if idx < len(out):
                    f = to_float(out[idx])
                    if f is not None:
                        out[idx] = f
        ws.append(out)
    n = len(rows) + 1
    if "result" in header and len(rows) > 0:
        col = get_column_letter(header.index("result") + 1)
        rng = f"A2:{get_column_letter(len(header))}{n}"
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'${col}2="win"'],
                        fill=PatternFill("solid", fgColor="C6EFCE")),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'${col}2="loss"'],
                        fill=PatternFill("solid", fgColor="FFC7CE")),
        )
    autosize(ws, len(header))


# ---------------- sheet 7: combined stats by bin ----------------

def build_combined_stats_sheet(ws, hist_header, hist_rows, matches_header, matches_rows):
    """Объединяет исторические ставки и новые подтверждённые матчи в одну
    разбивку по бинам кэфа. Источник: реальные ставки (история) + потенциальные
    ставки из matches (по first_cf_le_threshold)."""

    # 1. собираем (cf, won, source) из исторических ставок
    bets = []
    if hist_header:
        h = {name: i for i, name in enumerate(hist_header)}
        for r in hist_rows:
            cf = to_float(r[h["odds"]]) if h["odds"] < len(r) else None
            res = r[h["result"]] if h["result"] < len(r) else ""
            if cf is None or res not in ("win", "loss"):
                continue
            bets.append((cf, res == "win", "history"))

    # 2. собираем из новых matches (по правилу first_cf_le_threshold)
    if matches_header:
        h = {name: i for i, name in enumerate(matches_header)}
        for r in matches_rows:
            winner = r[h["winner"]] if h["winner"] < len(r) else ""
            if winner not in ("A", "B"):
                continue
            a_first = to_float(r[h["first_cf_le_threshold_a"]]) if h["first_cf_le_threshold_a"] < len(r) else None
            b_first = to_float(r[h["first_cf_le_threshold_b"]]) if h["first_cf_le_threshold_b"] < len(r) else None
            a_first_at = r[h["first_cf_le_threshold_a_at"]] if h["first_cf_le_threshold_a_at"] < len(r) else ""
            b_first_at = r[h["first_cf_le_threshold_b_at"]] if h["first_cf_le_threshold_b_at"] < len(r) else ""
            bet_on = bet_cf = None
            if a_first is not None and b_first is not None:
                bet_on = "A" if a_first_at <= b_first_at else "B"
                bet_cf = a_first if bet_on == "A" else b_first
            elif a_first is not None:
                bet_on, bet_cf = "A", a_first
            elif b_first is not None:
                bet_on, bet_cf = "B", b_first
            if bet_on is None:
                continue
            bets.append((bet_cf, bet_on == winner, "new"))

    if not bets:
        ws["A1"] = "Нет данных ни в historical_bets.csv ни в matches.csv"
        return

    bins = [
        ("<1.05", 0.0, 1.05),
        ("1.05-1.10", 1.05, 1.10),
        ("1.10-1.15", 1.10, 1.15),
        ("1.15-1.20", 1.15, 1.20),
        ("1.20-1.25", 1.20, 1.25),
        ("1.25-1.30", 1.25, 1.30),
        ("1.30-1.40", 1.30, 1.40),
        ("1.40-1.50", 1.40, 1.50),
        ("1.50-1.70", 1.50, 1.70),
        (">=1.70", 1.70, 99.0),
    ]
    headers = ["Бин кэфа", "Всего ставок", "Победы", "Поражения",
               "Винрейт %", "Средний кэф", "EV %", "Kelly %",
               "Из истории", "Из новых", "Комментарий"]
    ws.append(headers)
    style_header(ws, len(headers))

    for label, lo, hi in bins:
        in_b = [b for b in bets if lo <= b[0] < hi]
        n = len(in_b)
        if n == 0:
            continue
        wins = sum(1 for b in in_b if b[1])
        losses = n - wins
        wr = wins / n
        avg = sum(b[0] for b in in_b) / n
        ev = (wr * (avg - 1) - (1 - wr)) * 100
        kelly = ((wr * (avg - 1) - (1 - wr)) / (avg - 1) * 100) if avg > 1 else 0
        from_hist = sum(1 for b in in_b if b[2] == "history")
        from_new = sum(1 for b in in_b if b[2] == "new")
        comment = ""
        if n < 10:
            comment = "мало данных"
        elif ev < 0:
            comment = "EV отрицательный"
        elif kelly > 0:
            comment = f"Kelly ≈{kelly:.1f}%"
        ws.append([label, n, wins, losses,
                   round(wr * 100, 1), round(avg, 4),
                   round(ev, 2), round(kelly, 2),
                   from_hist, from_new, comment])

    # ИТОГО
    total = len(bets)
    total_wins = sum(1 for b in bets if b[1])
    ws.append([])
    ws.append(["ИТОГО", total, total_wins, total - total_wins,
               round(total_wins / total * 100, 1) if total else "",
               "", "", "",
               sum(1 for b in bets if b[2] == "history"),
               sum(1 for b in bets if b[2] == "new"),
               ""])
    last = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row=last, column=c).font = Font(bold=True)

    autosize(ws, len(headers), min_w=14)


# ---------------- sheet 8: threshold strategies ----------------

def build_threshold_sheet(ws, matches_header, matches_rows):
    """Для каждого порога (1.25, 1.20, 1.15, 1.10, 1.05) считает:
    сколько матчей дошло до этого кэфа, WR, EV, Kelly. Логика ставки:
    ставим на ту команду которая ПЕРВОЙ пробила этот порог."""
    if not matches_header:
        ws["A1"] = "matches.csv не найден"
        return

    h = {n: i for i, n in enumerate(matches_header)}
    thresholds = ("125", "120", "115", "110", "105")

    headers = ["Порог cf", "Матчей дошло", "Из них с winner", "Wins", "Losses",
               "WR %", "EV %", "Kelly %", "Quarter Kelly %", "Eighth Kelly %",
               "Совет по % банка"]
    ws.append(headers)
    style_header(ws, len(headers))

    for thr in thresholds:
        cf_thr = float(thr) / 100.0
        col_a = f"first_cf_le_{thr}_a"
        col_a_at = f"first_cf_le_{thr}_a_at"
        col_b = f"first_cf_le_{thr}_b"
        col_b_at = f"first_cf_le_{thr}_b_at"
        if col_a not in h:
            continue  # старый формат CSV без новых колонок

        bets = []  # (cf_at_entry, won)
        reached = 0
        with_winner = 0
        for r in matches_rows:
            if len(r) < len(matches_header):
                continue
            a_cf = to_float(r[h[col_a]])
            b_cf = to_float(r[h[col_b]])
            a_at = r[h[col_a_at]]
            b_at = r[h[col_b_at]]

            # хотя бы одна команда пробила?
            if a_cf is None and b_cf is None:
                continue
            reached += 1

            winner = r[h["winner"]]
            if winner not in ("A", "B"):
                continue
            with_winner += 1

            # ставим на ту что первой пробила
            bet_on = bet_cf = None
            if a_cf is not None and b_cf is not None:
                bet_on = "A" if a_at <= b_at else "B"
                bet_cf = a_cf if bet_on == "A" else b_cf
            elif a_cf is not None:
                bet_on, bet_cf = "A", a_cf
            else:
                bet_on, bet_cf = "B", b_cf

            bets.append((bet_cf, bet_on == winner))

        if not bets:
            ws.append([f"≤{cf_thr:.2f}", reached, with_winner, 0, 0, "", "", "", "", "", "нет данных"])
            continue

        n = len(bets)
        wins = sum(1 for b in bets if b[1])
        wr = wins / n
        avg_cf = sum(b[0] for b in bets) / n
        ev = (wr * (avg_cf - 1) - (1 - wr)) * 100
        kelly = ((wr * (avg_cf - 1) - (1 - wr)) / (avg_cf - 1) * 100) if avg_cf > 1 else 0
        q_kelly = kelly / 4
        e_kelly = kelly / 8

        if n < 10:
            advice = f"мало данных (n={n})"
        elif ev < 0:
            advice = "EV отрицательный — НЕ ставить"
        elif e_kelly > 10:
            advice = "5-8% банка"
        elif e_kelly > 5:
            advice = "3-5% банка"
        else:
            advice = "не более 2-3% банка"

        ws.append([
            f"≤{cf_thr:.2f}", reached, with_winner, wins, n - wins,
            round(wr * 100, 1), round(ev, 2), round(kelly, 2),
            round(q_kelly, 2), round(e_kelly, 2), advice,
        ])

    autosize(ws, len(headers), min_w=14)


# ---------------- main ----------------

def main():
    log_h, log_r = read_csv(LOG_FILE)
    m_h, m_r = read_csv(MATCHES_FILE)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Tick log"
    build_log_sheet(ws1, log_h, log_r)

    ws2 = wb.create_sheet("Matches")
    build_matches_sheet(ws2, m_h, m_r)

    ws3 = wb.create_sheet("Stats by bin")
    build_stats_sheet(ws3, m_h, m_r)

    ws4 = wb.create_sheet("Comeback hunt")
    build_comeback_sheet(ws4, m_h, m_r)

    ws5 = wb.create_sheet("Score vs odds")
    build_scatter_sheet(ws5, log_h, log_r)

    hist_h, hist_r = read_csv(HISTORICAL_FILE)
    ws6 = wb.create_sheet("Historical bets")
    build_historical_sheet(ws6, hist_h, hist_r)

    ws7 = wb.create_sheet("Combined stats")
    build_combined_stats_sheet(ws7, hist_h, hist_r, m_h, m_r)

    ws8 = wb.create_sheet("Threshold strategies")
    build_threshold_sheet(ws8, m_h, m_r)

    try:
        wb.save(OUT_FILE)
        print(f"OK -> {OUT_FILE}")
    except PermissionError:
        import time
        alt = OUT_FILE.replace(".xlsx", f"_{int(time.time())}.xlsx")
        wb.save(alt)
        print(f"[!] {OUT_FILE} занят (открыт в Excel?). Сохранил в: {alt}")
        print(f"   Закрой Excel и переименуй или запусти скрипт заново.")
    print(f"  Tick log:      {len(log_r)} строк")
    print(f"  Matches:       {len(m_r)} матчей")
    print(f"  Historical:    {len(hist_r)} ставок")


if __name__ == "__main__":
    main()
